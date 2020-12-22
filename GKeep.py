

#Script that returns checkmarks from Google Keep, inputs data to Excel sheet and resets all of the notes to their original state

#@author: Deividas Ovsianikovas


#------All imports----------
import gkeepapi
from openpyxl.descriptors.base import DateTime
import xlwt
import openpyxl
from openpyxl import Workbook
from xlutils.copy import copy
from xlrd import open_workbook
from datetime import datetime
from datetime import date
import os
from os import path
#------End of imports-------


#Init varialbes
columns = 1
total = 0 #Total num of tasks to be done
completed = 0 #Actual num of completed tasks today
maxCheckLists = 0
current_datetime = datetime.now()

WBfilename = 'StaffRecords.xlsx'
pathToFile = os.getcwd() + '\\' + WBfilename 


if(path.exists(pathToFile)):
    print("Path Exists, continue script")
else:
    wb = Workbook()
    wb.save(filename = WBfilename)


#------Google Keep Setup--------

keep = gkeepapi.Keep()
#!Important...make sure to delete login credentials....
success = keep.login('<Email>','<Password>')
gnotes = keep.all()

#------End of Google Keep Setup--------




#---Setting up the Excel Sheet------
rb = openpyxl.load_workbook(pathToFile) 
r_sheet = rb.active     

#Copy current sheet values to variable
wb = rb
ws = wb.active   

totalRows = ws.max_row #All existing rows in sheet
#----End Excel Sheet setup--------





#Get the max number of items in any note
for y in gnotes:
    if maxCheckLists < len(y.items):
        maxCheckLists = len(y.items)



#Shift all existing columns down based on the maximum items, leave gap of 2
for x in range(maxCheckLists + 6):
    ws.insert_rows(0)
    


#Loops going through all notes and all boxes within the gKeep notes
for y in gnotes:
    totalRows = 1
    ws.cell(row = 1, column = columns).value = y.title
    ws.cell(row = 2, column = columns).value = current_datetime.strftime('%a %y-%m-%d ')
    
    #print note values, for debugging
    #print (y.title)
    #print (y.text)

    for x in y.items:
            ws.cell(row = totalRows+2, column = columns).value = x.text
            ws.cell(row = totalRows+2, column = columns+1).value = x.checked

            if x.checked:
                completed+=1
            
            x.checked = False  
            totalRows+=1 

    total =  len(y.items)
    #Write values to the worksheet
    ws.cell(row = maxCheckLists+4, column = columns).value = "Total Completed: "
    ws.cell(row = maxCheckLists+4, column = columns+1).value = str(completed) + '/' + str(total)
    completed = 0

    #2 values for every note to leave a space between each
    columns += 3
    #end of loop





#Save Changes in keep and workbook
wb.save(WBfilename)
keep.sync()